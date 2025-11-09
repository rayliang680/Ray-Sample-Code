import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import os
from io import BytesIO

def analyze_growth_phases(time_data, value_data):
    """Analyze the four phases of microbial growth and return the time range of each phase"""
    # Calculate the slope (growth rate) between adjacent data points
    slopes = []
    for i in range(len(time_data) - 1):
        delta_y = value_data[i + 1] - value_data[i]
        delta_x = time_data[i + 1] - time_data[i]
        if delta_x != 0:
            slopes.append(delta_y / delta_x)
    
    if not slopes:
        return None
    
    # Calculate slope statistics (for threshold setting)
    avg_slope = np.mean(slopes)
    max_slope = np.max(slopes)
    min_slope = np.min(slopes)
    
    # Threshold setting (consistent with original VBS logic)
    lag_threshold = avg_slope * 0.3
    log_threshold = avg_slope + (max_slope - avg_slope) * 0.5
    death_threshold = avg_slope * 0.1
    
    # Initialize phase time variables
    lag_start = time_data[0] 
# if time_data else None
    lag_end = None
    log_start = None
    log_end = None
    stationary_start = None
    stationary_end = None
    death_start = None
    death_end = None
    
    # Phase flags
    in_lag = True
    in_log = False
    in_stationary = False
    in_death = False
    
    # Traverse slopes to identify growth phases
    for i in range(len(slopes)):
        if in_lag:
            # Lag phase → Log phase: slope exceeds log phase threshold for the first time
            if slopes[i] >= log_threshold:
                lag_end = time_data[i]
                log_start = time_data[i + 1]
                in_lag = False
                in_log = True
        elif in_log:
            # Log phase → Stationary phase: slope is below log phase threshold and above death phase threshold for 2 consecutive times
            if i >= 1 and slopes[i] < log_threshold and slopes[i-1] < log_threshold and slopes[i] > death_threshold:
                log_end = time_data[i]
                stationary_start = time_data[i + 1]
                in_log = False
                in_stationary = True
        elif in_stationary:
            # Stationary phase → Death phase: slope is below death phase threshold for the first time
            if slopes[i] <= death_threshold:
                stationary_end = time_data[i]
                death_start = time_data[i + 1]
                in_stationary = False
                in_death = True
                break  # Exit loop after finding death phase
    
    # Handle incompletely identified phases (use the last time point as the end)
    if in_stationary:
        stationary_end = time_data[-1]
    if in_death:
        death_end = time_data[-1]
    if lag_end is None:  # Log phase not identified
        lag_end = time_data[-1]
    
    return {
        "lag_phase": (lag_start, lag_end),
        "log_phase": (log_start, log_end),
        "stationary_phase": (stationary_start, stationary_end),
        "death_phase": (death_start, death_end),
        "slopes": slopes,
        "avg_slope": avg_slope,
        "max_slope": max_slope
    }

def generate_phase_report(ws, phase_data, title, start_row=1):
    """Generate four-phase analysis report in the worksheet"""
    ws[f'A{start_row}'] = title
    ws[f'A{start_row}'].font = ws[f'A{start_row}'].font.copy(bold=True, size=14)
    
    phases = [
        ("Lag Phase", phase_data["lag_phase"]),
        ("Log Phase", phase_data["log_phase"]),
        ("Stationary Phase", phase_data["stationary_phase"]),
        ("Death Phase", phase_data["death_phase"])
    ]
    
    row = start_row + 2
    ws[f'A{row-1}'] = "Growth Phase"
    ws[f'B{row-1}'] = "Time Range (hr)"
    ws[f'A{row-1}'].font = ws[f'A{row-1}'].font.copy(bold=True)
    ws[f'B{row-1}'].font = ws[f'B{row-1}'].font.copy(bold=True)
    
    for phase_name, (start, end) in phases:
        ws[f'A{row}'] = phase_name
        if start is not None and end is not None:
            ws[f'B{row}'] = f"{start:.2f} - {end:.2f}"
        elif start is not None:
            ws[f'B{row}'] = f"{start:.2f} - Ongoing"
        else:
            ws[f'B{row}'] = "Not Identified"
        row += 1
    
    # Add key parameters
    row += 2
    ws[f'A{row}'] = "Key Kinetic Parameters"
    ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True)
    row += 1
    ws[f'A{row}'] = "Average Growth Rate"
    ws[f'B{row}'] = f"{phase_data['avg_slope']:.4f} Units/hr"
    row += 1
    ws[f'A{row}'] = "Maximum Growth Rate"
    ws[f'B{row}'] = f"{phase_data['max_slope']:.4f} Units/hr"

def generate_comparison_report(ws, ypd_data, auxin_data):
    """Generate comparison report of two datasets"""
    ws['A1'] = "Comparison Conclusions of Microbial Growth Kinetics Between Two Groups"
    ws['A1'].font = ws['A1'].font.copy(bold=True, size=14)
    
    row = 3
    ws[f'A{row}'] = "Comparison Dimension"
    ws[f'B{row}'] = "GFP vector YPD"
    ws[f'C{row}'] = "GFP vector auxin"
    ws[f'D{row}'] = "Advantage Judgment"
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{row}'].font = ws[f'{col}{row}'].font.copy(bold=True)
    row += 1
    
    # Lag phase comparison
    ws[f'A{row}'] = "Lag Phase Duration (hr)"
    ypd_lag = ypd_data['lag_phase'][1] - ypd_data['lag_phase'][0]
    auxin_lag = auxin_data['lag_phase'][1] - auxin_data['lag_phase'][0]
    ws[f'B{row}'] = f"{ypd_lag:.2f}"
    ws[f'C{row}'] = f"{auxin_lag:.2f}"
    ws[f'D{row}'] = "YPD" if ypd_lag < auxin_lag else "auxin"
    row += 1
    
    # Log phase rate comparison
    ws[f'A{row}'] = "Maximum Growth Rate"
    ws[f'B{row}'] = f"{ypd_data['max_slope']:.4f}"
    ws[f'C{row}'] = f"{auxin_data['max_slope']:.4f}"
    ws[f'D{row}'] = "YPD" if ypd_data['max_slope'] > auxin_data['max_slope'] else "auxin"
    row += 1
    
    # Stationary phase duration comparison
    ws[f'A{row}'] = "Stationary Phase Duration (hr)"
    ypd_stationary = ypd_data['stationary_phase'][1] - ypd_data['stationary_phase'][0] if ypd_data['stationary_phase'][1] and ypd_data['stationary_phase'][0] else 0
    auxin_stationary = auxin_data['stationary_phase'][1] - auxin_data['stationary_phase'][0] if auxin_data['stationary_phase'][1] and auxin_data['stationary_phase'][0] else 0
    ws[f'B{row}'] = f"{ypd_stationary:.2f}" if ypd_stationary else "Not Identified"
    ws[f'C{row}'] = f"{auxin_stationary:.2f}" if auxin_stationary else "Not Identified"
    ws[f'D{row}'] = "YPD" if ypd_stationary > auxin_stationary else "auxin" if auxin_stationary else "None"
    row += 1
    
    # Overall growth capacity comparison
    ws[f'A{row}'] = "Overall Growth Capacity"
    ypd_growth = max(ypd_data['slopes']) if ypd_data['slopes'] else 0
    auxin_growth = max(auxin_data['slopes']) if auxin_data['slopes'] else 0
    ws[f'B{row}'] = "Strong" if ypd_growth > auxin_growth else "Weak"
    ws[f'C{row}'] = "Strong" if auxin_growth > ypd_growth else "Weak"
    ws[f'D{row}'] = "YPD" if ypd_growth > auxin_growth else "auxin"
    row += 1

def generate_analysis_report(ws, ypd_data, auxin_data, df):
    """Generate comprehensive analysis report"""
    ws['A1'] = "Comprehensive Analysis Report of Microbial Growth Kinetics"
    ws['A1'].font = ws['A1'].font.copy(bold=True, size=14)
    
    row = 3
    ws[f'A{row}'] = "I. Data Overview"
    ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True, size=12)
    row += 1
    ws[f'A{row}'] = f"The data contains {len(df)} time points, recording microbial growth data under two conditions:"
    row += 1
    ws[f'A{row}'] = "1. GFP vector YPD: GFP vector expression data in YPD medium"
    row += 1
    ws[f'A{row}'] = "2. GFP vector auxin: GFP vector expression data in medium containing auxin"
    row += 2
    
    ws[f'A{row}'] = "II. Growth Phase Analysis"
    ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True, size=12)
    row += 1
    ws[f'A{row}'] = "1. Growth phase characteristics of GFP vector YPD:"
    row += 1
    ws[f'A{row}'] = f"- Short lag phase, approximately {ypd_data['lag_phase'][1] - ypd_data['lag_phase'][0]:.2f} hr"
    row += 1
    ws[f'A{row}'] = f"- Rapid growth in log phase, with a maximum rate of {ypd_data['max_slope']:.4f} Units/hr"
    row += 1
    row += 1
    
    ws[f'A{row}'] = "2. Growth phase characteristics of GFP vector auxin:"
    row += 1
    ws[f'A{row}'] = f"- Lag phase of {auxin_data['lag_phase'][1] - auxin_data['lag_phase'][0]:.2f} hr"
    row += 1
    ws[f'A{row}'] = f"- Low overall growth rate, with a maximum rate of {auxin_data['max_slope']:.4f} Units/hr"
    row += 2
    
    ws[f'A{row}'] = "III. Key Findings"
    ws[f'A{row}'].font = ws[f'A{row}'].font.copy(bold=True, size=12)
    row += 1
    ws[f'A{row}'] = "1. YPD medium is more suitable for the rapid growth of the microorganism, showing higher growth rate and more distinct growth phases"
    row += 1
    ws[f'A{row}'] = "2. Auxin may have an inhibitory effect on the growth of the microorganism, leading to a significant decrease in growth rate"
    row += 1
    ws[f'A{row}'] = "3. The growth kinetic characteristics under the two conditions are significantly different, indicating that the medium composition has a great impact on the growth of the microorganism"

def create_charts(df, ws):
    """Create three visual charts and insert into the worksheet"""
    plt.rcParams["font.family"] = ["Arial", "Helvetica", "DejaVu Sans"]
    plt.rcParams['figure.dpi'] = 150
    
    # Chart 1: YPD time-dependent curve
    fig1, ax1 = plt.subplots(figsize=(8, 5))
    ax1.plot(df['Time [hr]'], df['GFP vector YPD'], marker='o', color='blue', label='GFP vector YPD')
    ax1.set_title('Microbial Growth Curve in YPD Medium')
    ax1.set_xlabel('Time (hr)')
    ax1.set_ylabel('Fluorescence Intensity')
    ax1.grid(True)
    ax1.legend()
    buf1 = BytesIO()
    fig1.savefig(buf1, format='png', bbox_inches='tight')
    buf1.seek(0)
    img1 = Image(buf1)
    ws.add_image(img1, 'A1')
    plt.close(fig1)
    
    # Chart 2: Auxin time-dependent curve
    fig2, ax2 = plt.subplots(figsize=(8, 5))
    ax2.plot(df['Time [hr]'], df['GFP vector auxin'], marker='s', color='red', label='GFP vector auxin')
    ax2.set_title('Microbial Growth Curve in Medium Containing Auxin')
    ax2.set_xlabel('Time (hr)')
    ax2.set_ylabel('Fluorescence Intensity')
    ax2.grid(True)
    ax2.legend()
    buf2 = BytesIO()
    fig2.savefig(buf2, format='png', bbox_inches='tight')
    buf2.seek(0)
    img2 = Image(buf2)
    ws.add_image(img2, 'A30')  # Place below the first chart
    plt.close(fig2)
    
    # Chart 3: Comparison curve of two datasets
    fig3, ax3 = plt.subplots(figsize=(8, 5))
    ax3.plot(df['Time [hr]'], df['GFP vector YPD'], marker='o', color='blue', label='YPD')
    ax3.plot(df['Time [hr]'], df['GFP vector auxin'], marker='s', color='red', label='auxin')
    ax3.set_title('Microbial Growth Comparison Between Two Media')
    ax3.set_xlabel('Time (hr)')
    ax3.set_ylabel('Fluorescence Intensity')
    ax3.grid(True)
    ax3.legend()
    buf3 = BytesIO()
    fig3.savefig(buf3, format='png', bbox_inches='tight')
    buf3.seek(0)
    img3 = Image(buf3)
    ws.add_image(img3, 'A60')  # Place below the second chart
    plt.close(fig3)

def process_excel(file_path):
    """Main function: Process Excel file and complete all requirements"""
    # Read data
    df = pd.read_excel(file_path)
    
    # Extract two datasets
    time_data = df['Time [hr]'].values
    ypd_data = df['GFP vector YPD'].values
    auxin_data = df['GFP vector auxin'].values
    
    # Analyze four phases
    ypd_phase = analyze_growth_phases(time_data, ypd_data)
    auxin_phase = analyze_growth_phases(time_data, auxin_data)
    
    # Load workbook
    wb = load_workbook(file_path)
    
    # 1. Create four-phase analysis report worksheets for two datasets
    if 'YPD Four-Phase Analysis' in wb.sheetnames:
        del wb['YPD Four-Phase Analysis']
    ypd_sheet = wb.create_sheet('YPD Four-Phase Analysis')
    generate_phase_report(ypd_sheet, ypd_phase, 'GFP vector YPD Growth Phase Analysis Report')
    
    if 'Auxin Four-Phase Analysis' in wb.sheetnames:
        del wb['Auxin Four-Phase Analysis']
    auxin_sheet = wb.create_sheet('Auxin Four-Phase Analysis')
    generate_phase_report(auxin_sheet, auxin_phase, 'GFP vector auxin Growth Phase Analysis Report')
    
    # 2. Create comparison conclusion worksheet
    if 'Comparison Conclusions' in wb.sheetnames:
        del wb['Comparison Conclusions']
    compare_sheet = wb.create_sheet('Comparison Conclusions')
    generate_comparison_report(compare_sheet, ypd_phase, auxin_phase)
    
    # 3. Create comprehensive analysis report worksheet
    if 'Comprehensive Analysis Report' in wb.sheetnames:
        del wb['Comprehensive Analysis Report']
    report_sheet = wb.create_sheet('Comprehensive Analysis Report')
    generate_analysis_report(report_sheet, ypd_phase, auxin_phase, df)
    
    # 4. Create visualization charts worksheet
    if 'Visualization Charts' in wb.sheetnames:
        del wb['Visualization Charts']
    chart_sheet = wb.create_sheet('Visualization Charts')
    create_charts(df, chart_sheet)
    
    # Save file
    output_path = os.path.splitext(file_path)[0] + '_Analysis_Results.xlsx'
    wb.save(output_path)
    return output_path

if __name__ == "__main__":
    # Replace with your Excel file path
    excel_file_path = "/Users/lshipeng/liangshipeng/test2.xlsx"  # Please modify to the actual file path
    result_file = process_excel(excel_file_path)
    print(f"Analysis completed! Results saved to: {result_file}")