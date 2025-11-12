# -*- coding: utf-8 -*-
from openpyxl import load_workbook

def macro1(ws):
    """Corresponds to the original VBS Macro 1: Determine if each row meets the conditions and mark "ok"/"not ok" in column A"""
    key_array1 = ["VWB", "PhiBT1", "pSAM2"]
    key_col1 = [15, 18, 21]  # Excel column numbers (1-based)
    key_array2 = ["phiC31", "phiC31"]
    key_col2 = [24, 27]
    key_string_col = 14  # Column 14 (column N) as the basis for judgment
    row_count = 11  # Start processing from row 11

    while ws.cell(row=row_count, column=key_string_col).value is not None:
        cell_value = str(ws.cell(row=row_count, column=key_string_col).value)
        is_ok1 = 0
        is_ok2 = 0
        is_find = 0

        # Process key_array1 logic
        for i in range(len(key_array1)):
            key = key_array1[i]
            col = key_col1[i]
            cell_val_col = ws.cell(row=row_count, column=col).value or 0  # Treat empty values as 0

            if key.lower() in cell_value.lower():  # Case-insensitive match
                if cell_val_col == 0:
                    is_ok1 = 0
                else:
                    is_ok1 = 1
                    break
            else:
                if cell_val_col == 0:
                    is_ok1 = 1
                    break
                else:
                    is_ok1 = 0

        # Process key_array2 logic (only when is_ok1=0)
        if is_ok1 == 0:
            for j in range(len(key_array2)):
                key = key_array2[j]
                col = key_col2[j]
                cell_val_col = ws.cell(row=row_count, column=col).value or 0

                if key.lower() in cell_value.lower():
                    if j == 0:
                        is_find = 1
                    if cell_val_col == 0:
                        if is_find == 1 and is_ok2 == 0:
                            is_ok2 = 0      
                            is_find = 1
                        else:
                            is_ok2 = 1
                            break
                    else:
                        is_ok2 = 1
                        break
                else:
                    if is_find == 1:
                        is_ok2 = 1
                        break
                    if cell_val_col == 1:
                        is_ok2 = 0
                        is_find = 0
                    else:
                        is_ok2 = 1
                        break

        # Write result to column A (column 1)
        if is_ok1 == 0 and is_ok2 == 0:
            ws.cell(row=row_count, column=1).value = "ok"
        else:
            ws.cell(row=row_count, column=1).value = "not ok"

        row_count += 1

def macro2(ws, wb):
    """Corresponds to the original VBS Macro 2: Calculate group accuracy rate and create a new worksheet to display results"""
    string_col = 14  # Column 14 (column N)
    key_string_col = 11  # Column 11 (column K) as the group identifier column
    row_count = 11
    last_string = ws.cell(row=row_count, column=key_string_col).value or ""

    # Initialize statistical arrays (corresponding to arr_key/arr_count/arr_count_ok in original VBS)
    arr_key = [""] * 10
    arr_count = [0] * 10
    arr_count_ok = [0] * 10
    if last_string:
        arr_key[0] = last_string  # Python list is 0-based, adapted to original logic

    while ws.cell(row=row_count, column=string_col).value is not None:
        is_find = 0
        current_key = ws.cell(row=row_count, column=key_string_col).value or ""

        # Detect group identifier changes, update last_string and add to array
        if current_key and current_key != last_string:
            last_string = current_key
            for i in range(len(arr_key)):
                if arr_key[i] == last_string:
                    is_find = 1
                    break
                elif arr_key[i] == "":
                    arr_key[i] = last_string
                    is_find = 1
                    break
            if not is_find:
                is_find = 1

        # Count total and correct numbers
        for i in range(len(arr_key)):
            if arr_key[i] == last_string:
                arr_count[i] += 1
                # Read judgment result from column A (column 1)
                if ws.cell(row=row_count, column=1).value == "ok":
                    arr_count_ok[i] += 1
                break

        row_count += 1

    # Create a new worksheet and write statistical results
    new_ws = wb.create_sheet(title="Statistical Results")
    # Write header
    new_ws.cell(row=2, column=1).value = "Parent Group Identifier"
    new_ws.cell(row=2, column=2).value = "Total Count"
    new_ws.cell(row=2, column=3).value = "Correct Count"
    new_ws.cell(row=2, column=4).value = "Accuracy Rate(%)"

    # Write statistical data
    for i in range(len(arr_key)):
        if arr_key[i]:
            row = i + 3  # Start writing data from row 3
            new_ws.cell(row=row, column=1).value = arr_key[i]
            new_ws.cell(row=row, column=2).value = arr_count[i]
            new_ws.cell(row=row, column=3).value = arr_count_ok[i]
            # Calculate accuracy rate (avoid division by zero)
            rate = (arr_count_ok[i] / arr_count[i] * 100) if arr_count[i] != 0 else 0
            new_ws.cell(row=row, column=4).value = round(rate, 2)  # Keep 2 decimal places

if __name__ == "__main__":
    # Please modify to your Excel file path (supports .xlsx format)
    excel_path = "Data_Analysis_Salnomycin.xlsx"
    # Load workbook (read_only=False allows writing)
    wb = load_workbook(excel_path)
    # Get the currently active worksheet (can be changed to a specified worksheet name, e.g., wb["Sheet1"])
    ws = wb.active

    # Execute the two macros
    macro1(ws)
    macro2(ws, wb)

    # Save the file (it is recommended to save as a new file to avoid overwriting the original file)
    output_path = "Salnomycin_Processed_File.xlsx"
    wb.save(output_path)
    wb.close()

    print(f"Processing completed! Results have been saved to: {output_path}")